#https://github.com/sferreyr/PythonPLCController

import os

from openpyxl import Workbook, load_workbook
import datetime

# NOMBRE ARCHIVO EXCEL
from openpyxl.styles import Font

hoy = str(datetime.date.today())
NAME_EXCEL = hoy + '.xlsx'
print(NAME_EXCEL)

class GenExcel:
    def __init__(self, DatosSepam, DatosGeneral):
        self.DatosSepam = DatosSepam
        self.DatosGeneral = DatosGeneral

    def Generar(self):
        general = False
        sepam = False
        if self.DatosGeneral != 0:
            general = True
            dicgeneral = self.DatosGeneral.__dict__
        if self.DatosSepam != 0:
            sepam = True
            dicsepam = self.DatosSepam.__dict__

        time = datetime.datetime.now()
        fileCheck = os.path.isfile(NAME_EXCEL)
        if fileCheck:
            wb = load_workbook(filename=NAME_EXCEL)
        else:
            wb = Workbook()
            print("No existe el archivo excel, se va a crear uno nuevo.")

        # Seteamos la hoja actual
        ws = wb.active

        # ESTILOS
        ft = Font(name='Arial', size=12)
        A1 = ws['A1']
        A1.font = ft
        A1 = "Datos Generales"

        C1 = ws['C1']
        C1.font = ft
        C1 = "Valores:"

        E1 = ws['E1']
        E1.font = ft
        E1 = "Fecha"

        if general == True:
            ws.append([A1, "", C1, "", E1])
            # Agregamos los atributos de GENERAL y sus valores.
            for key in dicgeneral:
                ws.append([key, "", dicgeneral[key], "", time])

        if sepam == True:
            ft = Font(name='Arial', size=12)
            A1 = ws['A1']
            A1.font = ft
            A1 = f"BOMBA N° {self.DatosSepam.NroBomba}"
            # Agregamos los atributos de SEPAM y sus valores.
            ws.append([A1])
            for key in dicsepam:
                ws.append([key, "", dicsepam[key], "", time])

        # Save the file
        wb.save(NAME_EXCEL)
